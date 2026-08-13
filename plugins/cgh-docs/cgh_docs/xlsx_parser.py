# -#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# __creation__ = 2026-07-29
# __author__ = "jndjama (Joy Ndjama)"
# __copyright__ = "Copyright 2026 ALTIKVA."
# __licence__ = "MIT & CC BY-NC-SA (https://www.altikva.com/licenses/LICENSE-1.0)"
# -#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# Description: XLSX parser: one section per sheet, with the header row and
#              the sheet size in the preview so column names are searchable.
#              Read-only mode, values only, so a big workbook costs little.

from __future__ import annotations

import logging
from pathlib import Path

from codegraph.plugin_api import BaseParser, FileIndex, SectionDef

_MAX_HEADER_CELLS = 30
_MAX_SCAN_CHARS = 200_000  # cap the extracted cell text fed to scanners
_log = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Best-effort xlsx to per-sheet sections. A corrupt workbook yields
    an empty index, never an exception."""

    lang = "xlsx"
    extensions = [".xlsx", ".xlsm"]
    extracts = ["sections"]
    description = "Excel sheets with their header row as document sections"

    def parse(self, path: Path) -> FileIndex:
        idx = FileIndex(path=str(path), lang=self.lang)
        wb = None
        try:
            import warnings

            import openpyxl

            with warnings.catch_warnings():
                # "Workbook contains no default style": cosmetic, and
                # it lands on the owner's stderr for every such file.
                warnings.simplefilter("ignore", UserWarning)
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            scan_parts: list[str] = []
            scan_len = 0
            for sheet_no, ws in enumerate(wb.worksheets, start=1):
                header: list[str] = []
                try:
                    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    header = [
                        str(c) for c in first[:_MAX_HEADER_CELLS] if c is not None
                    ]
                except StopIteration:
                    pass
                except Exception:
                    pass
                # Collect the cell values (not just the header) so PII in
                # data rows is scannable. Bounded across the whole workbook;
                # a giant sheet contributes until the budget is spent.
                if scan_len < _MAX_SCAN_CHARS:
                    scan_parts.append(f"# {ws.title}")
                    try:
                        for row in ws.iter_rows(values_only=True):
                            cells = [str(c) for c in row if c is not None]
                            if not cells:
                                continue
                            line = "\t".join(cells)
                            scan_parts.append(line)
                            scan_len += len(line) + 1
                            if scan_len >= _MAX_SCAN_CHARS:
                                break
                    except Exception:
                        pass
                preview = f"{ws.max_row or 0} rows x {ws.max_column or 0} cols" + (
                    f" | columns: {', '.join(header)}" if header else ""
                )
                idx.sections.append(
                    SectionDef(
                        id=f"{path}::sheet{sheet_no}",
                        title=ws.title,
                        level=1,
                        file_path=str(path),
                        start_line=sheet_no,
                        end_line=sheet_no,
                        body_preview=preview[:400],
                        anchor=f"sheet-{sheet_no}",
                    )
                )
            idx.scan_text = "\n".join(scan_parts)[:_MAX_SCAN_CHARS]
        except Exception:
            _log.warning("xlsx parse failed, indexed empty: %s", path)
            return idx
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
        return idx
